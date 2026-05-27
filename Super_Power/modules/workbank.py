import pandas as pd
import re
from rapidfuzz import fuzz
from itertools import combinations
import difflib
import tkinter as tk
from tkinter import simpledialog, messagebox, filedialog
import os
import shutil
from openpyxl import load_workbook

# =========================
# ROOT (Tkinter safe init)
# =========================
root = tk.Tk()
root.withdraw()

# =========================
# STRING CLEANING
# =========================
def clean_string(s):
    if pd.isna(s):
        return ""
    s = re.sub(r'\s+', ' ', str(s)).strip()
    s = s.replace('\u00A0', ' ')
    return s.lower()

# =========================
# DIFF HIGHLIGHT
# =========================
def highlight_diff(s1, s2):
    seq = difflib.SequenceMatcher(None, s1, s2)
    out = []

    for tag, i1, i2, j1, j2 in seq.get_opcodes():
        if tag == "equal":
            out.append(s1[i1:i2])
        elif tag in ("replace", "delete"):
            out.append(f"[{s1[i1:i2]}]")
        elif tag == "insert":
            out.append(f"(+{s2[j1:j2]})")

    return "".join(out)

# =========================
# SIMILARITY DETECTOR
# =========================
def detect_similar_lines_unique(strings, threshold=80):
    unique = list(dict.fromkeys(strings))
    pairs = []

    for s1, s2 in combinations(unique, 2):
        score = fuzz.ratio(s1, s2)
        if threshold <= score < 100:
            pairs.append((s1, s2, score, highlight_diff(s1, s2)))

    return pairs

# =========================
# SEGMENT BUILDER
# =========================
SIM_THRESHOLD = 90

def build_segment(row, prefix, pm_value, schemetitle_value):
    line = str(row.get("Line", ""))
    section = str(row.get("Section", ""))
    circuit = str(row.get("Circuit", ""))

    sim1 = fuzz.token_set_ratio(str(schemetitle_value), line)
    scheme_part = "" if sim1 >= SIM_THRESHOLD else f"{schemetitle_value} - "

    sim2 = fuzz.token_set_ratio(line, section)
    section_part = "" if sim2 >= SIM_THRESHOLD else f" - {section}"

    return f"{prefix}{pm_value} - {scheme_part}{circuit} - {line}{section_part}"

# =========================
# GUI CHOICE
# =========================
def choose_string_gui(s1, s2, diff, sheet_name, similarity):
    msg = (
        f"Sheet: {sheet_name}\n"
        f"Similarity: {similarity}%\n\n"
        f"1: {s1}\n2: {s2}\n\n"
        f"Diff:\n{diff}\n\n"
        f"Type 1, 2, or keep_both:"
    )

    choice = simpledialog.askstring(
        "Choose Line",
        msg,
        parent=root
    )

    if choice is None:
        return [s1]

    choice = choice.strip().lower()

    if choice == "1":
        return [s1]
    elif choice == "2":
        return [s2]
    elif choice == "keep_both":
        return [s1, s2]
    else:
        messagebox.showinfo("Invalid", "Defaulting to first option")
        return [s1]

# =========================
# PROCESS SHEET
# =========================
def process_target_sheet(tp_path, sheet, pm_value, schemetitle_value, threshold=90):

    usecols = [1, 2, 3, 4, 6, 8, 9, 11, 13]

    df = pd.read_excel(
        tp_path,
        sheet_name=sheet,
        header=3,
        usecols=usecols,
        engine="openpyxl"
    )

    df.columns = [
        "Circuit", "Section", "Line", "ENID", "Item",
        "Work Description", "Span", "Quantity", "Comment"
    ]

    df["Line"] = df["Line"].astype(str).apply(clean_string)

    similar = detect_similar_lines_unique(df["Line"].tolist(), threshold)

    for s1, s2, score, diff in similar:
        chosen = choose_string_gui(s1, s2, diff, sheet, score)

        if len(chosen) == 1:
            df.loc[df["Line"].isin([s1, s2]), "Line"] = chosen[0]

    # remove rows with too many missing
    missing = df.apply(
        lambda r: r.astype(str).str.contains("missing", case=False, na=False).sum(),
        axis=1
    )
    df = df[missing < 2]

    df["Segment"] = df.apply(
        lambda r: build_segment(r, "M - " if sheet == "Pole Materials" else "C - ",
                                pm_value, schemetitle_value),
        axis=1
    )

    return df

# =========================
# FILE SELECTION
# =========================
control_path = filedialog.askopenfilename(
    title="Select Control File",
    filetypes=[("Excel files", "*.xlsx *.xlsm")]
)

tp_path = filedialog.askopenfilename(
    title="Select Target Price File",
    filetypes=[("Excel files", "*.xlsx *.xlsm")]
)

if not control_path or not tp_path:
    print("Missing files")
    exit()

# =========================
# READ METADATA
# =========================
wb_tp = load_workbook(tp_path, data_only=True)
ws = wb_tp["Target"]

pm_value = ws["D21"].value or "Unknown"
schemetitle_value = ws["D4"].value or "Unknown"
PID = ws["D22"].value or ""
wb_tp.close()

# =========================
# PROCESS SHEETS
# =========================
tp_sheets = ["Pole Materials", "OHL"]
tp_data = []

for sh in tp_sheets:
    try:
        df = process_target_sheet(tp_path, sh, pm_value, schemetitle_value)
        tp_data.append(df)
    except Exception as e:
        print(f"Error {sh}: {e}")

if not tp_data:
    print("No data processed")
    exit()

tp_combined = pd.concat(tp_data, ignore_index=True)

# =========================
# COPY CONTROL FILE
# =========================
base, ext = os.path.splitext(control_path)
updated_path = f"{base}_{schemetitle_value}{ext}"
shutil.copy2(control_path, updated_path)

wb = load_workbook(updated_path, keep_vba=True)

# =========================
# WRITE AUX
# =========================
ws_aux = wb["Auxiliar"]

start = 4
for i, row in tp_combined.iterrows():
    ws_aux.cell(start + i, 1, row["Segment"])
    ws_aux.cell(start + i, 2, row["ENID"])
    ws_aux.cell(start + i, 3, row["Work Description"])
    ws_aux.cell(start + i, 4, row["Quantity"])
    ws_aux.cell(start + i, 5, row["Span"])
    ws_aux.cell(start + i, 6, row["Item"])
    ws_aux.cell(start + i, 7, row["Comment"])

# =========================
# SAVE
# =========================
wb.save(updated_path)
wb.close()

print(f"Done → {updated_path}")
