import pandas as pd
import re
from rapidfuzz import fuzz
from itertools import combinations
import difflib
import tkinter as tk
from tkinter import simpledialog, messagebox, filedialog
import os
import shutil
from openpyxl import load_workbook

# Hide main Tk window but keep it for dialogs
root = tk.Tk()
root.withdraw()

# ---------------------------
# --- String cleaning function ---
# ---------------------------
def clean_string(s):
    if pd.isna(s):
        return ""
    s = re.sub(r'\s+', ' ', str(s)).strip()
    s = s.replace('\u00A0', ' ')
    return s.lower()

# ---------------------------
# --- Highlight differences ---
# ---------------------------
def highlight_diff(s1, s2, mode="change"):
    if mode == "keep_both":
        return {"original_1": s1, "original_2": s2}

    seq = difflib.SequenceMatcher(None, s1, s2)
    highlighted = []

    for tag, i1, i2, j1, j2 in seq.get_opcodes():
        if tag == "equal":
            highlighted.append(s1[i1:i2])
        elif tag in ("replace", "delete"):
            highlighted.append(f"[{s1[i1:i2]}]")
        elif tag == "insert":
            highlighted.append(f"(+{s2[j1:j2]})")
    return "".join(highlighted)

# ---------------------------
# --- Detect similar lines ---
# ---------------------------
def detect_similar_lines_unique(strings, threshold=80):
    """
    Returns list of tuples:
    (s1, s2, similarity, diff)
    """
    unique_strings = list(dict.fromkeys(strings))  # Only unique strings
    similar_pairs = []

    for s1, s2 in combinations(unique_strings, 2):
        similarity = fuzz.ratio(s1, s2)
        if similarity >= threshold and similarity < 100:
            diff = highlight_diff(s1, s2)
            similar_pairs.append((s1, s2, similarity, diff))

    return similar_pairs


def choose_string_gui(s1, s2, diff, sheet_name, similarity, parent_root):
    """
    Returns a LIST of chosen lines:
    - [s1] if 1
    - [s2] if 2
    - [s1, s2] if keep_both
    """
    dialog = tk.Toplevel(parent_root)
    dialog.withdraw()
    dialog.transient(parent_root)
    dialog.grab_set()

    msg = (
        f"Sheet: {sheet_name}\n"
        f"Similarity: {similarity}%\n\n"
        f"1: {s1}\n2: {s2}\n\n"
        f"Diff:\n{diff}\n\n"
        f"Type 1, 2, or 'keep_both':"
    )

    chosen = simpledialog.askstring(
        title="Choose Line",
        prompt=msg,
        initialvalue="1",
        parent=dialog
    )

    result = [s1]  # default
    if chosen is not None:
        choice = chosen.strip().lower()
        if choice == "1":
            result = [s1]
        elif choice == "2":
            result = [s2]
        elif choice == "keep_both":
            result = [s1, s2]
        else:
            messagebox.showinfo(
                "Invalid choice",
                f"No valid input entered. Using first line by default:\n{s1}",
                parent=dialog
            )

    dialog.grab_release()
    dialog.destroy()
    return result

# ---------------------------
# --- GUI function for choosing string ---
# ---------------------------
def process_target_sheet(tp_path, sheet, pm_value, schemetitle_value, root, threshold=90):
    """
    Reads a sheet from Target Price, detects similar lines, allows user choice,
    keeps all original rows intact, and updates the Segment column correctly.
    """
    usecols = [1, 2, 3, 4, 6, 8, 9, 11, 13]  # B,D,E,G,I,J,L,N
    df = pd.read_excel(tp_path, sheet_name=sheet, header=3, usecols=usecols, engine="openpyxl")
    df.columns = [
        "Circuit", "Section", "Line", "ENID", "Item",
        "Work Description", "Span", "Quantity", "Comment"
    ]

    # Clean "Line"
    df["Line"] = df["Line"].astype(str).apply(clean_string)

    # Detect similar lines
    similar = detect_similar_lines_unique(df["Line"].tolist(), threshold=threshold)

    if similar:
        print(f"⚠️ Similar lines detected in '{sheet}':")
        for s1, s2, score, diff in similar:
            print(f"[{score}% similarity]\n  1: {s1}\n  2: {s2}\n  Diff: {diff}\n")
            chosen_lines = choose_string_gui(s1, s2, diff, sheet_name=sheet, similarity=score, parent_root=root)

            if len(chosen_lines) == 1:
                # Replace all occurrences of s1 and s2 with the chosen line
                df.loc[df["Line"].isin([s1, s2]), "Line"] = chosen_lines[0]

            # If "keep_both", do nothing, keep all rows as they are
            # All rows remain intact and segments will be updated later

    else:
        print(f"✅ No similar lines detected in '{sheet}'.")

    # Skip rows with "Missing" in 2 or more columns
    missing_counts = df.apply(lambda row: row.astype(str).str.contains("missing", case=False, na=False).sum(), axis=1)
    df = df[missing_counts < 2]

    # Add Segment based on current Line (after possible renaming)
    prefix = "M - " if sheet == "Pole Materials" else "C - "
    df["Segment"] = df.apply(build_segment, axis=1, prefix=prefix, pm_value=pm_value, schemetitle_value=schemetitle_value)

    print(f"✅ Processed {len(df)} rows from '{sheet}'")
    return df
# ---------------------------
# --- Build segment string ---
# ---------------------------
SIM_THRESHOLD = 90

def build_segment(row, prefix, pm_value, schemetitle_value):
    similarity = fuzz.token_set_ratio(str(schemetitle_value), str(row["Line"]))
    scheme_part = "" if similarity >= SIM_THRESHOLD else f"{schemetitle_value} - "

    line_section_similarity = fuzz.token_set_ratio(str(row["Line"]), str(row["Section"]))
    section_part = "" if line_section_similarity >= SIM_THRESHOLD else f" - {row['Section']}"

    return f"{prefix}{pm_value} - {scheme_part}{row['Circuit']} - {row['Line']}{section_part}"

# --- Select Control file ---
control_path = filedialog.askopenfilename(
    title="Hello Sir, please select the Template_CF Excel",
    filetypes=[("Excel files", "*.xlsx *.xlsm")]
)
if not control_path:
    print("❌ No Control File selected.")
    exit()
print(f"\n📘 Selected Control file: {control_path}")

# --- Select Target Price file ---
tp_path = filedialog.askopenfilename(
    title="Hello again Sir, please select the Target Price Excel file you wish to format (.xlsm)",
    filetypes=[("Excel Macro-Enabled Workbook", "*.xlsm"), ("All files", "*.*")]
)
if not tp_path:
    print("❌ No Target Price file selected.")
    exit()
print(f"📗 Selected Target Price file: {tp_path}")

# --- Read PM value from Target Price sheet 'Target', cell D21 ---
try:
    wb_tp = load_workbook(tp_path, data_only=True)
    ws_target = wb_tp["Target"]

    pm_value = ws_target["D21"].value or "Unknown"
    schemetitle_value = ws_target["D4"].value or "Unknown"
    PID = ws_target["D22"].value or ""

    wb_tp.close()
    print(f"🔹 PM value read from Target Price (Target!D21): {pm_value}")
    print(f"🔹 Scheme Title read from Target Price (Target!D4): {schemetitle_value}")

except Exception as e:
    print(f"⚠️ Could not read values from Target sheet: {e}")
    pm_value = "Unknown"
    schemetitle_value = "Unknown"
    PID = ""

# --- Read Target Price data from sheets ---
tp_sheets = ["Pole Materials", "OHL"]
usecols = [1, 2, 3, 4, 6, 8, 9, 11, 13]  # B,D,E,G,I,J,L,N

tp_data = []
for sheet in tp_sheets:
    try:
        df = process_target_sheet(tp_path, sheet, pm_value, schemetitle_value, root)
        tp_data.append(df)
    except Exception as e:
        print(f"⚠️ Error reading '{sheet}': {e}")

# Combine sheets
tp_combined = pd.concat(tp_data, ignore_index=True)


# --- Ask for custom suffix for output filename ---
suffix = schemetitle_value

# --- Copy Control file to preserve formulas ---
base, ext = os.path.splitext(control_path)
updated_path = f"{base}_{suffix}{ext}"
shutil.copy2(control_path, updated_path)
print(f"📂 Copied Control file → {updated_path}")

# --- Open copied Control file ---
wb = load_workbook(updated_path, keep_vba=True)

# --- Write to 'Auxiliar' sheet ---
ws_aux = wb["Auxiliar"]
first_empty_aux = next((row for row in range(4, ws_aux.max_row + 2) if ws_aux.cell(row=row, column=1).value in (None, "")), ws_aux.max_row + 1)

for i, (_, row) in enumerate(tp_combined.iterrows()):
    ws_aux.cell(row=first_empty_aux + i, column=1, value=row["Segment"])
    ws_aux.cell(row=first_empty_aux + i, column=2, value=row["ENID"])
    ws_aux.cell(row=first_empty_aux + i, column=3, value=row["Work Description"])
    ws_aux.cell(row=first_empty_aux + i, column=4, value=row["Quantity"])
    ws_aux.cell(row=first_empty_aux + i, column=5, value=row["Span"])
    ws_aux.cell(row=first_empty_aux + i, column=6, value=row["Item"])
    ws_aux.cell(row=first_empty_aux + i, column=7, value=row["Comment"])

# --- Write to 'Block1' sheet ---
block_segments = []
if "Block1" in wb.sheetnames:
    ws_block = wb["Block1"]
    first_empty_block = next((row for row in range(30, ws_block.max_row + 2) if ws_block.cell(row=row, column=1).value in (None, "")), ws_block.max_row + 1)

    for i, (_, row) in enumerate(tp_combined.iterrows()):
        target_row = first_empty_block + i
        ws_block.cell(row=target_row, column=1, value=row["Segment"])
        ws_block.cell(row=target_row, column=2, value=row["ENID"])
        ws_block.cell(row=target_row, column=3, value=row["Work Description"])
        ws_block.cell(row=target_row, column=4, value=row["Quantity"])
        ws_block.cell(row=target_row, column=6, value=row["Quantity"])
        ws_block.cell(row=target_row, column=20, value=row["Span"])
        ws_block.cell(row=target_row, column=21, value=row["Comment"])
        ws_block.cell(row=target_row, column=91, value=PID)
        block_segments.append(row["Segment"])
else:
    print("⚠️ Sheet 'Block1' not found in Control file!")

# --- Write to 'PA CONTROL' sheet ---
if "PA CONTROL" in wb.sheetnames:
    ws_pa = wb["PA CONTROL"]
    first_empty_pa = next((row for row in range(3, ws_pa.max_row + 2) if ws_pa.cell(row=row, column=1).value in (None, "")), ws_pa.max_row + 1)

    unique_segments = sorted(set(s for s in block_segments if s and str(s).strip()))
    for i, seg in enumerate(unique_segments):
        target_row = first_empty_pa + i
        ws_pa.cell(row=target_row, column=1, value=seg)
        ws_pa.cell(row=target_row, column=2, value="MAT" if str(seg).startswith("M") else "CONST")

else:
    print("⚠️ Sheet 'PA CONTROL' not found in Control file!")

# --- Save workbook ---
wb.save(updated_path)
wb.close()
print(f"💾 Data successfully transferred to '{updated_path}' with PM='{pm_value}' and all formulas intact!")
